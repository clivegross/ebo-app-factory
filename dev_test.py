# Read the test HTML file
from ebo_app_factory.html_file_builder import EBOHTMLFileBuilder
import xml.etree.ElementTree as ET
from ebo_app_factory.xml_app_factory import (
    ApplicationTemplate,
    FactoryInputsFromSpreadsheet,
    ApplicationFactory,
)
import os
from openpyxl import load_workbook


make_stream_monitoring_interfaces = False
make_stream_html = True

max_items_per_file = 300

with open("tests/data/{{streamname}}.html", "r", encoding="utf-8") as f:
    template_html = f.read()

placeholder = "{{streamname}}"
output_xml = "tests/data/streams_html_ebo.xml"

item_workbook_path = "tests/data/Streams.xlsx"


# function to load NAME column from Excel file
def load_names_from_excel(file_path):
    """Load the NAME column from the default sheet of an Excel file."""
    workbook = load_workbook(file_path)
    sheet = workbook.active  # Use the default/active sheet

    # Find the NAME column
    name_column_index = None
    for col_idx, cell in enumerate(sheet[1], 1):  # First row (header)
        if cell.value == "NAME":
            name_column_index = col_idx
            break

    if name_column_index is None:
        raise ValueError("No 'NAME' column found in the Excel file")

    # Extract all values from the NAME column (excluding header)
    names = []
    for row in sheet.iter_rows(
        min_row=2, min_col=name_column_index, max_col=name_column_index
    ):
        cell_value = row[0].value
        if cell_value is not None:  # Skip empty cells
            names.append(str(cell_value))

    return names


# function to replace placeholders in the HTML content
def replace_placeholders(html_content, placeholders):
    result = html_content
    for key, value in placeholders.items():
        result = result.replace(key, value)
    return result


def create_html_files(template_html, placeholder, streamnames, output_xml):
    builder = EBOHTMLFileBuilder(ebo_version="6.0.4.90")
    for streamname in streamnames:
        # Define the placeholders and their replacements
        placeholders = {
            placeholder: streamname,
        }
        html = replace_placeholders(template_html, placeholders)
        print(
            f"HTML content after replacing placeholders:\n{html[:200]}..."
        )  # Print first 200 chars for brevity

        # Create HTML file object
        html_obj, object_type = builder.create_and_add_html_file(
            name=streamname,
            html_content=html,
            # description="{{Description}}",
            # note1="{{Note1}}",
            # note2="{{Note2}}",
        )
    # Write the XML to a file
    builder.write_xml(output_xml)


if make_stream_html:
    # Load streamnames from Excel file
    try:
        streamnames = load_names_from_excel(item_workbook_path)
        print(
            f"Loaded {len(streamnames)} stream names from Excel: {streamnames[:5]}{'...' if len(streamnames) > 5 else ''}"
        )
    except Exception as e:
        print(f"Error loading from Excel: {e}")

    # Create HTML files for each streamname
    create_html_files(template_html, placeholder, streamnames, output_xml)


if make_stream_monitoring_interfaces:
    # declare filenames/paths and settings here:
    item_workbook_path = "tests/data/stream_items.xlsx"
    template_path = "tests/data/Camera Stream Monitoring {{streamname}}.xml"
    output_path = "tests/data/stream_monitoring_apps.xml"
    # max_items_per_file = 6
    # set item_workbook sheetnames and corresponding app templates:
    app_template = ApplicationTemplate(template_path)
    factory_inputs = FactoryInputsFromSpreadsheet(xlfile=item_workbook_path)

    # Act: Create the XML
    app_factory = ApplicationFactory(
        template_child_elements_dict=app_template.template_child_elements_dict,
        factory_placeholders=factory_inputs.factory_placeholders,
        factory_copy_substrings=factory_inputs.factory_copy_substrings,
        xml_out_file=output_path,
    )
    app_factory.make_document(max_items_per_file=max_items_per_file)
