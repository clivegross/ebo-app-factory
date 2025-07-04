import xml
import openpyxl
from xml.dom import minidom
import xml.etree.ElementTree as ET
import sys
import re
import os

from ebo_app_factory.ebo_xml_builder import EBOXMLBuilder
from .xmlutils import (
    convert_minidom_to_etree,
    extract_mustache_tags_from_xml,
    find_and_clean_folder_elements,
)

# <?xml version="1.0" encoding="UTF-8"?>
# <ObjectSet ExportMode="Special" Note="TypesFirst" Version="3.2.1.630">
#   <MetaInformation>
#     <ExportMode Value="Special"/>
#     <RuntimeVersion Value="3.2.1.630"/>
#     <SourceVersion Value="3.2.1.630"/>
#     <ServerFullPath Value="/ebo_app_factory"/>
#   </MetaInformation>
#   <Types>
# 	 	...
# 	</Types>
# 	<ExportedObjects>
# 		<OI NAME="VAV-L21-INT4" TYPE="system.base.Folder">
# 			...
# 		</OI>
# 	</ExportedObjects>
# </ObjectSet>


class ApplicationTemplate(object):

    def __init__(self, xml_in_file, print_result=False):
        """
        creates a dictionary of lists of child DOM Elements from the template xml file
        where each key represents the tagname of each eleemnt that should live in the root element of the DOM

        ApplicationTemplate.template_child_elements_dict = {
                'Types': [<DOM Element: ObjectType at 0x7f5665bec700>, <DOM Element: ObjectType at 0x7f5665adef70>],
                'ExportedObjects': [<DOM Element: OI at 0x7f5665afe430>]
        }
        """
        self.xml_in_doc = minidom.parse(xml_in_file)
        # list of nodes to export children
        self.template_nodes = ["Types", "ExportedObjects"]
        self.template_child_elements_dict = self.get_template_child_elements_dict()
        if print_result:
            for key, value in self.template_child_elements_dict.items():
                for element in value:
                    print(key, value)
                    print(element.toxml())

    def get_child_nodes_by_element_tagname(self, tagname, elements_only=False):
        """
        assumes 1st element returned in getElementsByTagName
        returns list of child nodes inside Element 'tagname'
        """
        nodes = self.xml_in_doc.getElementsByTagName(tagname)

        if not nodes:  # If the list is empty
            return []  # Return an empty list

        if elements_only:
            return self.get_child_elements(nodes[0].childNodes)
        else:
            return nodes[0].childNodes

    def get_attr_if_exists(self, node, attr_name):
        if node.hasAttribute(attr_name):
            return node.getAttribute(attr_name)
        else:
            return None

    def get_child_elements(self, node):
        """
        returns list of child nodes of type Node.ELEMENT_NODE only
        """
        return [child for child in node if child.nodeType == minidom.Node.ELEMENT_NODE]

    def get_template_child_elements_dict(self):
        """
        returns dictionary of lists of xml child element nodes required for xml app AppTemplate.
        Keys are found in self.template_nodes: Types, ExportedObjects
        """
        template_child_elements_dict = {}
        for key in self.template_nodes:
            template_child_elements_dict[key] = self.get_child_nodes_by_element_tagname(
                key, elements_only=True
            )
        return template_child_elements_dict


class FactoryInputsFromSpreadsheet(object):

    def __init__(self, xlfile=None, sheetname=None, print_result=False):
        """
        read in a spreadsheet containing a tables of:
        - app template placeholder substrings
        - app copy replacement substrings
        By default, all sheets are read except 'meta'. To limit which sheets are read, set sheetname to a string mathcing the sheet name or a list of strings equal to each sheet name required.
        each non empty cell of the first row of each sheet read is stored as a template placeholder substring key:value store
        each non empty second and subsequent row represents an app to be copied from the template, stored as a list of key:value stores,
        the cells of which correspond to replacement strings

        Example:

        Sheet1
        x		A		B		C
        1		VAV-1	Zn1		Room 2.31
        2		VAV-2	Zn2		Meeting Room 7
        3		VAV-3.2	Zn3B	Level 3 reception

        factory_placeholders = {
        'Sheet1A': 'VAV-1',
        'Sheet1B': 'Zn1',
        'Sheet1C': 'Room 2.31'
        }

        factory_placeholders_sorted = {
                'Sheet1': {
                        'Sheet1A': 'VAV-1',
                        'Sheet1B': 'Zn1',
                        'Sheet1C': 'Room 2.31'
                }
        }

        factory_copy_substrings = [
                {
                        'Sheet1A': 'VAV-2',
                        'Sheet1B': 'Zn2',
                        'Sheet1C': 'Meeting Room 7'
                },
                {
                        'Sheet1A': 'VAV-3.2',
                        'Sheet1B': 'Zn3B',
                        'Sheet1C': 'Level 3 reception'
                },
        ]

        factory_copy_substrings_sorted = {
                'Sheet1': [
                        {
                                'Sheet1A': 'VAV-2',
                                'Sheet1B': 'Zn2',
                                'Sheet1C': 'Meeting Room 7'
                        },
                        {
                                'Sheet1A': 'VAV-3.2',
                                'Sheet1B': 'Zn3B',
                                'Sheet1C': 'Level 3 reception'
                        },
                ]
        }

        """
        self.xlfile = xlfile
        self.show_progress = True
        self.create_factory_inputs_from_excel(sheetname=sheetname)
        if print_result:
            print(self.factory_placeholders)
            print(self.factory_copy_substrings)

    def create_factory_inputs_from_excel(self, sheetname=None):
        """
        if sheetname not specified, all sheets will be read into one big list (except 'meta')
        if sheetname is str, only read the sheet with name sheetname
        if sheetname is list, read each member of list as a sheetname

        self.factory_copy_substrings is a flattened list of copy substrings
        factory_copy_substrings_sorted is a dict of lists of copy substrings,
        where each key represents a sheet
        """

        workbook = openpyxl.load_workbook(self.xlfile, data_only=True)

        if sheetname == None:
            allsheetnames = workbook.sheetnames
            sheetnames = [s for s in allsheetnames if "meta" not in s]
        elif isinstance(sheetname, str):
            sheetnames = [sheetname]
        elif isinstance(sheetname, list):
            sheetnames = sheetname

        self.factory_placeholders = {}
        self.factory_placeholders_sorted = {}
        self.factory_copy_substrings = []
        self.factory_copy_substrings_sorted = {}

        if self.show_progress:
            print("\nCreating factory inputs from:", sheetnames)

        for sheetname in sheetnames:
            (placeholders, factory_copy_substrings) = (
                self.create_factory_inputs_from_xl_sheet(sheetname, workbook)
            )
            self.factory_placeholders.update(placeholders)
            self.factory_copy_substrings.extend(factory_copy_substrings)
            self.factory_placeholders_sorted[sheetname] = placeholders
            self.factory_copy_substrings_sorted[sheetname] = factory_copy_substrings

    def create_factory_inputs_from_xl_sheet(self, sheetname, workbook):

        sheet = workbook[sheetname]
        placeholders = {}
        factory_copy_substrings = []
        for row in sheet.iter_rows():
            factory_copy = {}
            first_row = True
            for cell in row:
                key = sheetname + cell.column_letter
                if cell.row == 1:
                    placeholders[key] = cell.value
                else:
                    first_row = False
                    factory_copy[key] = cell.value
            if not first_row:
                factory_copy_substrings.append(factory_copy)
        return (placeholders, factory_copy_substrings)


class ApplicationFactory(object):

    def __init__(
        self,
        template_child_elements_dict=None,
        factory_placeholders=None,
        factory_copy_substrings=None,
        xml_out_file=None,
        ebo_version="4.0.1.86",
        ebo_server_full_path="/EBOApplicationFactory_v0.1",
        ebo_export_mode="Special",
        show_progress=True,
    ):
        self.show_progress = show_progress
        self.xml_out_file = xml_out_file
        self.template_child_elements_dict = template_child_elements_dict
        self.factory_placeholders = factory_placeholders
        self.factory_copy_substrings = factory_copy_substrings
        self.xml_builder = EBOXMLBuilder(
            ebo_version=ebo_version,
            server_full_path=ebo_server_full_path,
            export_mode=ebo_export_mode,
        )

    def stdout_progress(self, step, total_steps):
        if self.show_progress:
            sys.stdout.write("\r")
            sys.stdout.write("%d%%" % (step / total_steps * 100))
            sys.stdout.flush()
        return step + 1

    def make_document(self, write_result=True, print_result=False):
        """
        The xml document is constructed as follows:
        <ObjectSet>
                {{ header stuff }}
                <Types>
                        {{ self.factory_copies_dict['Types'] }}
                </Types>
                <ExportedObjects>
                        {{ self.factory_copies_dict['ExportedObjects'] }}
                </ExportedObjects>
        </ObjectSet>
        """
        # check if factory copies has already been created
        if not hasattr(self, "factory_copies_dict"):
            self.make_copies()
        # report progress
        if self.show_progress:
            print("\nCreating document...")
        size = 0
        for node in self.factory_copies_dict:
            size += len(self.factory_copies_dict[node])
        progress = 1
        for item in self.factory_copies_dict["Types"]:
            # Convert minidom Element to ElementTree Element
            etree_element = convert_minidom_to_etree(item)
            self.xml_builder.add_object_type(etree_element)
        for item in self.factory_copies_dict["ExportedObjects"]:
            # Convert minidom Element to ElementTree Element
            # Check if it's a minidom Element before converting
            if isinstance(item, xml.dom.minidom.Element):
                etree_element = convert_minidom_to_etree(item)
            elif isinstance(item, ET.Element):
                etree_element = item  # Already an ElementTree Element
            else:
                # Handle other types or raise an error
                raise TypeError(f"Unexpected element type: {type(item)}")
            self.xml_builder.add_to_exported_objects(etree_element)
            # report progress
            progress = self.stdout_progress(progress, size)
        # loop through dictionary keys for each element to insert children
        # for node, elements in self.factory_copies_dict.items():
        #     # create an empty child element inside root DOM Element ObjectSet
        #     factory_element = self.factory_doc.createElement(node)
        #     self.factory_doc.documentElement.appendChild(factory_element)
        #     # self.doc_template.getElementsByTagName(self.doc_root_element_tagname)[0].appendChild()
        #     # loop through elements and insert as children
        #     for child_element in elements:
        #         self.factory_doc.getElementsByTagName(node)[0].appendChild(
        #             child_element
        #         )
        if print_result:
            # print(self.factory_doc.toprettyxml(encoding="utf-8"))
            # print(self.factory_doc)
            print(self.xml_builder.to_pretty_xml())
        if write_result:
            if self.show_progress:
                print('\nWriting document to "' + self.xml_out_file + '" ...')
                self.xml_builder.write_xml(self.xml_out_file)
            # with open(self.xml_out_file, "wb") as outfile:
            #     outfile.write(self.factory_doc.toxml(encoding="utf-8"))
            if self.show_progress:
                print("\nDone.\n")

    def make_copies(self):
        """
        self.template_child_elements_dict AND factory_copies_dict = {
                'Types': [
                        <DOM Element: ObjectType at 0x7fac26249f70>,
                        <DOM Element: ObjectType at 0x7fac26144ee0>
                ],
                'ExportedObjects': [
                        <DOM Element: OI at 0x7fac261653a0>
                ]
        }
        self.factory_copy_substrings = [
                {'Sheet1A': 'VAV-L16-INT1', 'Sheet1B': 'L16-INT1'},
                {'Sheet1A': 'VAV-L16-INT10', 'Sheet1B': 'L16-INT10'},
                {'Sheet1A': 'VAV-L16-INT11', 'Sheet1B': 'L16-INT11'}
        ]
        """
        # report progress
        if self.show_progress:
            print("Creating copies...")
        size = len(self.factory_copy_substrings)
        progress = 1
        # create empty factory copies dictionary
        factory_copies_dict = {key: [] for key in self.template_child_elements_dict}

        # Copy Types elements as-is (no modifications needed)
        if "Types" in self.template_child_elements_dict:
            factory_copies_dict["Types"] = self.template_child_elements_dict["Types"][:]

        # Only make copies for ExportedObjects, not Types
        if "ExportedObjects" in self.template_child_elements_dict:
            elements = self.template_child_elements_dict["ExportedObjects"]
            # loop through copy strings list
            for copy_substrings in self.factory_copy_substrings:
                for element in elements:
                    copy_element = self.replace_placeholders(element, copy_substrings)
                    factory_copies_dict["ExportedObjects"].append(copy_element)
                # report progress
                progress = self.stdout_progress(progress, size)
        self.factory_copies_dict = factory_copies_dict

    def get_unique_copies_for_placeholder(self, placeholder):
        """
        returns a list of unique copy strings for the placeholder
        for the placeholder value, make a list of unique copies strings
        if self.factory_placeholders: dict = {'controllersA': 'Controller 1', 'controllersB': 'IRD-ICG-B03'}
        and self.factory_copy_substrings: list = [{'controllersA': 'Controller 2', 'controllersB': 'IRD-ICG-B04'},
                                                  {'controllersA': 'Controller 3', 'controllersB': 'IRD-ICG-B04'},
                                                  {'controllersA': 'Controller 4', 'controllersB': 'IRD-ICG-B05'},
                                                  {'controllersA': 'Controller 5', 'controllersB': 'IRD-ICG-L00'},
                                                  {'controllersA': 'Controller 6', 'controllersB': 'IRD-ICG-L00'}}
        and placeholder = 'IRD-ICG-B03'
        then return {'controllersB': ['IRD-ICG-B04', 'IRD-ICG-B05', 'IRD-ICG-L00']}
        """
        unique_copies = {}

        # Find which key has the placeholder value
        for key, placeholder_value in self.factory_placeholders.items():
            if placeholder_value == placeholder:
                # Collect unique values for this key from factory_copy_substrings
                unique_values = set()
                for copy_dict in self.factory_copy_substrings:
                    if key in copy_dict:
                        unique_values.add(copy_dict[key])

                # Convert set to list and store
                unique_copies[key] = list(unique_values)

        return unique_copies

    def make_copies_in_folders(self, placeholder_folder_name):
        """
        based on the placeholder_folder_name, create an EBO folder for each unique copy
        and put all copied with mtching value for this column inside the folder

        """
        # report progress
        if self.show_progress:
            print("Creating copies in folders...")
        size = len(self.factory_copy_substrings)
        progress = 1
        factory_copies_dict = {key: [] for key in self.template_child_elements_dict}

        # Copy Types elements as-is (no modifications needed)
        if "Types" in self.template_child_elements_dict:
            factory_copies_dict["Types"] = self.template_child_elements_dict["Types"][:]

        copy_folder_names = self.get_unique_copies_for_placeholder(
            placeholder_folder_name
        )
        print(copy_folder_names)
        # get key of placeholder_folder_name
        # if self.factory_placeholders: dict = {'controllersA': 'Controller 1', 'controllersB': 'IRD-ICG-B03'} and placeholder_folder_name = 'IRD-ICG-B03' return 'controllersB'
        folder_key = None
        for key, value in self.factory_placeholders.items():
            if value == placeholder_folder_name:
                folder_key = key
                break
        print(f"Folder key for '{placeholder_folder_name}': {folder_key}")
        folders = []
        for folder_name in copy_folder_names[folder_key]:
            # filter self.factory_copy_substrings to only those that have the folder_key: copy_folder_name
            filtered_copy_substrings = [
                item
                for item in self.factory_copy_substrings
                if item.get(folder_key) == folder_name
            ]
            # print(f"Folder name for '{folder_name}': {filtered_copy_substrings}")
            folder_element = self.xml_builder.create_folder(folder_name)

            elements = self.template_child_elements_dict["ExportedObjects"]
            # loop through copy strings list
            for copy_substrings in filtered_copy_substrings:
                for element in elements:
                    copy_element = self.replace_placeholders(element, copy_substrings)
                    etree_element = convert_minidom_to_etree(copy_element)
                    folder_element.append(etree_element)
                # report progress
            progress = self.stdout_progress(progress, size)
            folders.append(folder_element)
        factory_copies_dict["ExportedObjects"] = folders
        self.factory_copies_dict = factory_copies_dict

    def replace_placeholders(self, element, copy_substrings):
        """
        find and replace xml element template placeholder strings with copy strings
        """
        # convert DOM Element to xml string
        factory_copy_element_str = element.toxml()
        # find and replace placeholders with copy values
        for key, placeholder_value in self.factory_placeholders.items():
            copy_value = copy_substrings.get(key)
            if copy_value is None:
                # print(f"Warning: copy_substrings[{key}] is None. Skipping replacement.")
                continue  # Skip this iteration if the value is None
            else:
                # Ensure the copy_value is a string
                copy_value = str(copy_value)
                factory_copy_element_str = factory_copy_element_str.replace(
                    placeholder_value, copy_value
                )

        # convert xml string back to DOM Element
        # print("convert xml string back to DOM Element:")
        # print_first_22_lines(factory_copy_element_str)
        # print(element.tagName)
        # handle ampersand, parse to xml
        factory_copy_element_str = re.sub(
            r"&(?!(?:amp|lt|gt|apos|quot);)", "&amp;", factory_copy_element_str
        )
        # factory_copy_element_str = escape(factory_copy_element_str, {"&": "&amp;"})
        factory_copy_element = minidom.parseString(
            factory_copy_element_str
        ).getElementsByTagName(element.tagName)[0]
        # self.xml_in_doc.getElementsByTagName(tagname)[0]
        return factory_copy_element


def print_first_22_lines(s):
    lines = s.splitlines()
    for line in lines[:22]:
        print(line)


class ApplicationFactoryManager(object):

    def __init__(
        self,
        template_map=None,
        xlfile=None,
        sheetname=None,
        xml_out_file_prefix=None,
        max_items_per_file=None,  # Add the optional argument for max_items_per_file
        ebo_version="4.0.1.86",
        ebo_server_full_path="/EBOApplicationFactory_v0.1",
        ebo_export_mode="Special",
        show_progress=True,
    ):
        self.show_progress = show_progress
        self.xlfile = xlfile
        self.xml_out_file_prefix = xml_out_file_prefix
        self.template_map = template_map
        self.max_items_per_file = (
            max_items_per_file  # Store the max_items_per_file value
        )

        self.get_factory_inputs(sheetname=sheetname)
        self.get_app_templates()

    def get_app_templates(self):
        if self.show_progress:
            print("\nCreating template documents...")
        for group, items in self.template_map.items():
            print('\nCreating template document for "' + group + '" applications...')
            print(items)
            items["elements"] = ApplicationTemplate(
                items["templateFilename"], print_result=False
            ).template_child_elements_dict

    def get_factory_inputs(self, sheetname=None):
        if self.show_progress:
            print('\nCreating factory inputs from workbook "' + self.xlfile + '"')
        self.factory_inputs = FactoryInputsFromSpreadsheet(
            self.xlfile, sheetname=sheetname, print_result=False
        )
        self.factory_placeholders_sorted = (
            self.factory_inputs.factory_placeholders_sorted
        )
        self.factory_copy_substrings_sorted = (
            self.factory_inputs.factory_copy_substrings_sorted
        )

    def make_documents(self):
        for (
            group,
            factory_copy_substrings,
        ) in self.factory_copy_substrings_sorted.items():

            if self.show_progress:
                print('\nStarting production on "' + group + '" applications...')

            # Calculate the number of files needed based on max_items_per_file
            if self.max_items_per_file is None:
                num_files = 1
            else:
                num_files = len(factory_copy_substrings) // self.max_items_per_file
                if len(factory_copy_substrings) % self.max_items_per_file != 0:
                    num_files += 1

            # Create multiple XML files if needed
            for i in range(num_files):
                # Calculate the start and end index for the current batch
                # handle self.max_items_per_file = None
                if self.max_items_per_file is None:
                    start_idx = 0
                    end_idx = len(factory_copy_substrings)
                else:
                    start_idx = i * self.max_items_per_file
                    end_idx = min(
                        (i + 1) * self.max_items_per_file, len(factory_copy_substrings)
                    )
                batch_copy_substrings = factory_copy_substrings[start_idx:end_idx]

                # Create ApplicationFactory for each batch of instances
                app_factory = ApplicationFactory(
                    template_child_elements_dict=self.template_map[group]["elements"],
                    factory_placeholders=self.factory_placeholders_sorted[group],
                    factory_copy_substrings=batch_copy_substrings,
                    xml_out_file=self.xml_out_file_prefix
                    + "_"
                    + group
                    + "_"
                    + str(i + 1)
                    + ".xml",  # Append index to file name
                )
                app_factory.make_document()

            # app_factory = ApplicationFactory(
            # 	template_child_elements_dict=self.template_map[group]['elements'],
            # 	factory_placeholders=self.factory_placeholders_sorted[group],
            # 	factory_copy_substrings=factory_copy_substrings,
            # 	xml_out_file=self.xml_out_file_prefix+'_'+group+'.xml',
            # )
            # app_factory.make_document()


def make_empty_factory_app_list_spreadsheet(xml_template_paths, xl_out_file=None):
    """
    creates a spreadsheet containing sheets:
    - app template placeholder mustache tags in first row (one sheet per template)
    - meta sheet with object list from all templates
    """
    # create a workbook and add a worksheet
    wb = openpyxl.Workbook()
    # Check if "Sheet1" exists
    for default_sheet in wb.sheetnames:
        # Get the sheet by name
        sheet = wb[default_sheet]
        # Remove "Sheet1"
        wb.remove(sheet)
    for i, template_path in enumerate(xml_template_paths):
        # get name of template file without extension or path
        template_name = os.path.splitext(os.path.basename(template_path))[0]
        # create a new sheet
        ws = wb.create_sheet(title=template_name)
        # get mustache tags from xml
        mustache_tags = extract_mustache_tags_from_xml(template_path)
        # add mustache "{{}}" braces around each tag
        mustache_tags = ["{{" + tag + "}}" for tag in mustache_tags]
        # insert mustahce tags into first row
        ws.append(mustache_tags)
    # add meta sheet with object list from all templates
    meta_ws = wb.create_sheet(title="meta")
    for i, template_path in enumerate(xml_template_paths):
        # now extract the object list from the xml
        elements = find_and_clean_folder_elements(template_path)
        # example elements:
        # elements = [{'NAME': 'ZnRh', 'TYPE': 'server.point.AV', 'DESCR': 'zone air humidity sensor', 'FOLDER': 'Variables'},
        # 	{'NAME': 'ZnTmp', 'TYPE': 'server.point.AV', 'DESCR': 'zone air temperature sensor', 'FOLDER': 'Variables'},
        # 	{'NAME': 'Air Filter Clogged', 'TYPE': 'alarm.ChangeOfStateAlarm', 'DESCR': 'Air Filter Clogged', 'FOLDER': 'Alarms'}]
        # if first template, write headers
        if i == 0:
            # extract the keys from the first element, assume all elements have the same keys
            keys = list(elements[0].keys())
            headers = ["TEMPLATE"] + keys
            meta_ws.append(headers)  # Write the headers to the first row
        # get name of template file without extension or path
        template_name = os.path.splitext(os.path.basename(template_path))[0]
        # Add the template name to the beginning of each element dictionary
        for element in elements:
            element.attrib["TEMPLATE"] = template_name
            # Create a row by extracting values from the element's attributes
            # Note: headers should include 'template' and other attribute names you want to extract
            row = [
                element.attrib.get(key, "") for key in headers
            ]  # Use .get to avoid KeyError if an attribute is missing
            meta_ws.append(row)
    # Save the workbook to the specified file path
    if xl_out_file:
        wb.save(xl_out_file)
    return wb


def template_folder_elements_list_of_lists(template_paths):
    """
    Processes elements from each template_path and returns them as a list of lists.

    Parameters:
    - template_paths (dict): A dictionary where keys are template names and values are lists of elements (each element is a dictionary of key-value pairs).

    Returns:
    - List[List[str]]: A list of lists, where each inner list represents a row with the template name and the values of KEY1, KEY2, KEY3, etc.
    """
    rows = []  # Initialize the list to hold all rows

    # Check if there are any template paths to process
    if template_paths:
        for i, template_path in enumerate(template_paths):
            # now extract the object list from the xml
            elements = find_and_clean_folder_elements(template_path)
            # extract the keys from the first element, assume all elements have the same keys
            keys = list(elements[0].keys())
            # get name of template file without extension or path
            template_name = os.path.splitext(os.path.basename(template_path))[0]
            # Add the template name to the beginning of each element dictionary
            for element in elements:
                element.insert(0, template_name)
            # Append the elements to the rows list
            rows.extend(elements)
    return rows

    # loop through xml_template_paths

    # save the workbook
    if xl_out_file:
        wb.save(xl_out_file)
    return wb


# EXECUTE
if __name__ == "__main__":

    ########################
    # Basic example
    ########################
    # declare filenames/paths here
    xl_in_file = "examples/basic apps example.xlsx"
    xml_in_file = "examples/VAV-L21-INT4 application special.xml"
    xml_out_file = "examples/generated_ebo_apps_basic_example.xml"

    # instantiate AppTemplate object object
    app_template = ApplicationTemplate(xml_in_file, print_result=False)
    # instantiate FactoryInputsFromSpreadsheet object
    factory_inputs = FactoryInputsFromSpreadsheet(xl_in_file, print_result=False)
    # instantiate ApplicationFactory object and make xml
    app_factory = ApplicationFactory(
        template_child_elements_dict=app_template.template_child_elements_dict,
        factory_placeholders=factory_inputs.factory_placeholders,
        factory_copy_substrings=factory_inputs.factory_copy_substrings,
        xml_out_file=xml_out_file,
    )
    # app_factory.make_copies()
    app_factory.make_document()

    ########################
    # Advanced example
    ########################
    # declare filenames/paths here
    xl_sorted_in_file = "examples/sorted apps example.xlsx"
    # create dictionary mapping Excel sheet names to template xml files
    template_map = {
        "L2-3-All3StgHtg": {
            "templateFilename": "examples/VAV-L21-NW2 application special.xml"
        },
        "L4-12-3StgHtg": {
            "templateFilename": "examples/VAV-L21-NW2 application special.xml"
        },
        "L13-15-3StgHtg": {
            "templateFilename": "examples/VAV-L21-NW2 application special.xml"
        },
        "L16-32-3StgHtg": {
            "templateFilename": "examples/VAV-L21-NW2 application special.xml"
        },
        "1StgHtg": {
            "templateFilename": "examples/VAV-L04-INT09 application special.xml"
        },
        "L2-3NoHtg": {
            "templateFilename": "examples/VAV-L21-INT4 application special.xml"
        },
        "L4-12NoHtg": {
            "templateFilename": "examples/VAV-L21-INT4 application special.xml"
        },
        "L13-15NoHtg": {
            "templateFilename": "examples/VAV-L21-INT4 application special.xml"
        },
        "L16-32NoHtg": {
            "templateFilename": "examples/VAV-L21-INT4 application special.xml"
        },
    }
    # instantiate ApplicationFactoryManager object
    app_factory_manager = ApplicationFactoryManager(
        template_map=template_map,
        xlfile=xl_sorted_in_file,
        xml_out_file_prefix="examples/example_ebo_apps",
    )
    # make xml files
    app_factory_manager.make_documents()
